import streamlit as st
from datetime import date
from calendar import monthrange

from config.settings import BANK_DISPLAY_NAMES, ESG_EMOJIS

MONTHS_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def render_month_nav() -> tuple[date, date]:
    today = date.today()

    if "nav_year" not in st.session_state:
        st.session_state.nav_year = today.year
    if "nav_month" not in st.session_state:
        st.session_state.nav_month = today.month

    year = st.session_state.nav_year
    month = st.session_state.nav_month

    at_min = (year == today.year and month == 1)
    at_max = (year == today.year and month == today.month)

    # Navigation rendered in main body
    col_prev, col_label, col_next = st.columns([1, 4, 1])

    with col_prev:
        if st.button("◀", disabled=at_min, use_container_width=True, key="nav_prev"):
            if month == 1:
                st.session_state.nav_month = 12
                st.session_state.nav_year = year - 1
            else:
                st.session_state.nav_month = month - 1
            st.rerun()

    with col_label:
        with st.popover(
            f"{MONTHS_PT[month - 1]} {year}",
            use_container_width=True,
        ):
            st.markdown("**Ir para mês/ano**")
            available_years = [today.year]
            sel_year = st.selectbox("Ano", available_years, index=0, key="pop_year")
            max_month = today.month if sel_year == today.year else 12
            sel_month = st.selectbox(
                "Mês",
                list(range(1, max_month + 1)),
                format_func=lambda m: MONTHS_PT[m - 1],
                index=max_month - 1,
                key="pop_month",
            )
            if st.button("Ir", use_container_width=True, key="pop_go"):
                st.session_state.nav_year = sel_year
                st.session_state.nav_month = sel_month
                st.rerun()

    with col_next:
        if st.button("▶", disabled=at_max, use_container_width=True, key="nav_next"):
            if month == 12:
                st.session_state.nav_month = 1
                st.session_state.nav_year = year + 1
            else:
                st.session_state.nav_month = month + 1
            st.rerun()

    month_start = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    month_end = date(year, month, last_day)

    return month_start, month_end


def render_sidebar_filters() -> tuple[list[str], list[str]]:
    st.sidebar.markdown(
        '<div style="padding:4px 0 12px;">'
        '<div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">'
        '<span style="font-size:1.6rem;line-height:1;">🌱</span>'
        '<span style="font-size:1.2rem;font-weight:800;">ESG News Calendar</span>'
        '</div>'
        '<div style="font-size:0.78rem;color:#6c757d;line-height:1.5;">'
        'Rastreamento diário de notícias ambientais, sociais e de governança dos principais bancos brasileiros.'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.sidebar.markdown(
        '<div class="sidebar-section-title">Bancos</div>',
        unsafe_allow_html=True,
    )

    all_banks = list(BANK_DISPLAY_NAMES.keys())
    all_bank_names = [BANK_DISPLAY_NAMES[b] for b in all_banks]

    selected_bank_names = st.sidebar.multiselect(
        "Bancos",
        options=all_bank_names,
        default=[],
        label_visibility="collapsed",
        placeholder="Todos os bancos",
    )
    selected_banks = [b for b in all_banks if BANK_DISPLAY_NAMES[b] in selected_bank_names]

    st.sidebar.markdown(
        '<div class="sidebar-section-title">Categoria ESG</div>',
        unsafe_allow_html=True,
    )

    esg_options = {"E": ("Ambiental", "#2d6a4f"), "S": ("Social", "#1d3557"), "G": ("Governança", "#6a0572")}
    selected_esg = []
    for tag, (label, color) in esg_options.items():
        emoji = ESG_EMOJIS[tag]
        checked = st.sidebar.checkbox(
            f"{emoji} {label}",
            value=True,
            key=f"esg_{tag}",
        )
        if checked:
            selected_esg.append(tag)

    return selected_banks, selected_esg


def render_export_csv(articles: list[dict], filename: str = "esg_news.xlsx") -> None:
    """Botão de download XLSX das notícias filtradas — formatado para Excel."""
    if not articles:
        return
    import io
    from datetime import date as _date
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from config.settings import BANK_DISPLAY_NAMES, ESG_LABELS

    def _fmt_date(v) -> str:
        if isinstance(v, _date):
            return v.strftime("%d/%m/%Y")
        if isinstance(v, str) and len(v) >= 10:
            try:
                return _date.fromisoformat(v[:10]).strftime("%d/%m/%Y")
            except Exception:
                return v
        return str(v) if v else ""

    def _clean(v) -> str:
        if v is None:
            return ""
        return str(v).replace("\r", " ").replace("\n", " ").strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Notícias ESG"

    headers = ["Data", "Banco", "Categoria ESG", "Título", "Resumo", "Fonte", "Link"]
    widths = [12, 18, 15, 55, 65, 22, 30]

    header_fill = PatternFill("solid", fgColor="1E5631")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 24

    wrap = Alignment(wrap_text=True, vertical="top")
    for a in articles:
        row = [
            _fmt_date(a.get("data")),
            BANK_DISPLAY_NAMES.get(a.get("banco_tag", ""), a.get("banco_tag", "")),
            ESG_LABELS.get(a.get("esg_tag", ""), a.get("esg_tag", "")),
            _clean(a.get("titulo")),
            _clean(a.get("resumo")),
            _clean(a.get("fonte")),
            _clean(a.get("link")),
        ]
        r = ws.max_row + 1
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.alignment = wrap
        # Link clicável
        link_cell = ws.cell(row=r, column=7)
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.font = Font(color="0563C1", underline="single")
            link_cell.value = "Abrir notícia"

    # Congela a primeira linha
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    st.sidebar.download_button(
        label="📥 Exportar Excel",
        data=buf.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


_COLOR_DOMAIN = ["Ambiental", "Social", "Governança"]
_COLOR_RANGE = ["#2ecc71", "#3498db", "#9b59b6"]


def _build_trend_df(monthly_counts: list[dict]):
    import pandas as pd
    ordered = sorted(monthly_counts, key=lambda m: m["month"])
    rows = []
    for m in ordered:
        try:
            y, mo = m["month"].split("-")
            label = f"{MONTHS_PT[int(mo) - 1][:3]}/{y[2:]}"
        except Exception:
            label = m["month"]
        rows.append({
            "mes": label,
            "Ambiental": m.get("E", 0),
            "Social": m.get("S", 0),
            "Governança": m.get("G", 0),
        })
    return pd.DataFrame(rows)


def render_trend_chart(monthly_counts: list[dict], window: int = 4) -> None:
    """Tendência ESG com dois modos: histograma (paginado) e linha do tempo (consolidado)."""
    if not monthly_counts:
        st.caption("Sem dados suficientes para o gráfico de tendência.")
        return

    df = _build_trend_df(monthly_counts)

    col_title, col_mode = st.columns([6, 3])
    with col_title:
        st.markdown("### 📈 Tendência ESG")
    with col_mode:
        mode = st.radio(
            "Visualização",
            options=["Histograma", "Linha do tempo"],
            horizontal=True,
            label_visibility="collapsed",
            key="trend_mode",
        )

    if mode == "Histograma":
        _render_histogram(df, window)
    else:
        _render_timeline(df)


def _render_histogram(df, window: int) -> None:
    total = len(df)
    n_pages = max(1, (total + window - 1) // window)

    if "trend_page" not in st.session_state:
        st.session_state.trend_page = n_pages - 1
    st.session_state.trend_page = max(0, min(st.session_state.trend_page, n_pages - 1))
    page = st.session_state.trend_page

    col_prev, col_label, col_next = st.columns([1, 6, 1])
    with col_prev:
        if st.button("◀", key="trend_prev", disabled=(page == 0), use_container_width=True):
            st.session_state.trend_page -= 1
            st.rerun()
    with col_label:
        st.markdown(
            f"<div style='text-align:center;padding-top:8px;font-size:0.85rem;opacity:0.75;'>"
            f"Página {page + 1}/{n_pages}</div>",
            unsafe_allow_html=True,
        )
    with col_next:
        if st.button("▶", key="trend_next", disabled=(page == n_pages - 1), use_container_width=True):
            st.session_state.trend_page += 1
            st.rerun()

    window_df = df.iloc[page * window:page * window + window]
    month_order = window_df["mes"].tolist()

    try:
        import altair as alt
        long = window_df.melt(id_vars="mes", var_name="Categoria", value_name="Notícias")
        chart = (
            alt.Chart(long)
            .mark_bar()
            .encode(
                x=alt.X("mes:N", title=None, sort=month_order,
                        axis=alt.Axis(labelAngle=0)),
                xOffset=alt.XOffset("Categoria:N"),
                y=alt.Y("Notícias:Q", title=None),
                color=alt.Color(
                    "Categoria:N",
                    scale=alt.Scale(domain=_COLOR_DOMAIN, range=_COLOR_RANGE),
                    legend=alt.Legend(orient="bottom", title=None),
                ),
                tooltip=["mes", "Categoria", "Notícias"],
            )
            .properties(height=300)
            .configure_view(strokeWidth=0)
        )
        st.altair_chart(chart, use_container_width=True, theme=None)
    except Exception:
        _render_trend_html(window_df.set_index("mes"))


def _render_timeline(df) -> None:
    """Linha do tempo consolidada com todos os meses disponíveis."""
    month_order = df["mes"].tolist()
    try:
        import altair as alt
        long = df.melt(id_vars="mes", var_name="Categoria", value_name="Notícias")
        chart = (
            alt.Chart(long)
            .mark_line(
                point=alt.OverlayMarkDef(size=70, filled=True),
                interpolate="monotone",
                strokeWidth=3,
            )
            .encode(
                x=alt.X("mes:N", title=None, sort=month_order,
                        axis=alt.Axis(labelAngle=0)),
                y=alt.Y("Notícias:Q", title=None),
                color=alt.Color(
                    "Categoria:N",
                    scale=alt.Scale(domain=_COLOR_DOMAIN, range=_COLOR_RANGE),
                    legend=alt.Legend(orient="bottom", title=None),
                ),
                tooltip=["mes", "Categoria", "Notícias"],
            )
            .properties(height=320)
            .configure_view(strokeWidth=0)
        )
        st.altair_chart(chart, use_container_width=True, theme=None)
    except Exception:
        _render_trend_html(df.set_index("mes"))


def _render_trend_html(df) -> None:
    """Renderização manual quando st.line_chart falha — grouped bar chart."""
    max_val = max(df.values.max(), 1)
    colors = {"Ambiental": "#2ecc71", "Social": "#3498db", "Governança": "#9b59b6"}
    BAR_MAX_HEIGHT = 120  # px

    # Legenda
    legend = " &nbsp;•&nbsp; ".join(
        f'<span style="color:{c};font-weight:600;">● {name}</span>'
        for name, c in colors.items()
    )

    # Grupos por mês, com 3 barras (E/S/G) lado a lado
    groups = []
    for idx in df.index:
        bars = []
        for col in df.columns:
            val = int(df.loc[idx, col])
            h = int(val / max_val * BAR_MAX_HEIGHT) if val > 0 else 2
            bars.append(
                f'<div style="display:flex;flex-direction:column;align-items:center;gap:4px;">'
                f'<div style="font-size:0.7rem;font-weight:700;color:{colors[col]};">{val}</div>'
                f'<div style="width:22px;height:{h}px;background:{colors[col]};border-radius:3px 3px 0 0;"></div>'
                f'</div>'
            )
        groups.append(
            f'<div style="display:flex;flex-direction:column;align-items:center;gap:6px;flex:1;">'
            f'<div style="display:flex;gap:6px;align-items:flex-end;height:{BAR_MAX_HEIGHT + 20}px;">{"".join(bars)}</div>'
            f'<div style="font-size:0.75rem;opacity:0.75;font-weight:500;">{idx}</div>'
            f'</div>'
        )

    html = (
        f'<div style="margin-bottom:8px;font-size:0.8rem;">{legend}</div>'
        f'<div style="display:flex;gap:16px;align-items:flex-end;padding:8px 0 16px 0;">'
        f'{"".join(groups)}'
        f'</div>'
    )
    st.markdown(html, unsafe_allow_html=True)
